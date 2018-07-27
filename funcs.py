import re
import os
from openpyxl import load_workbook
from nltk import sent_tokenize, word_tokenize


def normalWord(word):
    # if re.search('s$',word):
    #     r = word[:len(word)-1]
    if not len(word):
        return ""
    if not word.isupper():
        word = word.lower()
    start = 0
    end = len(word) - 1
    while word[start] == " " and start < (len(word) - 1):
        start += 1
    while word[end] == " " and end > 0:
        end -= 1
    return word[start: end+1]


def isEnglish(src):
    for d in src:
        if ord(d)>126:
            return False
    return True


def correctStart(word):
    a = ["-", "*", "+", "=", "/", "^", "#", "@", "~" ,"•","· ", "?"]
    if word[0] in a and len(word) == 1:
        return False
    result = word
    for i in range(len(word)):
        if word[i] in a:
            result = word[i+1:]
        else:
            return normalWord(result)
    return False


def script_pre(script ):
    return script.replace("/", " ")


def check_contain_chinese(str):
    count = 0
    for ch in str:
        if u'\u4e00' <= ch <= u'\u9fff':
            count+=1
    if count:
        return count
    return False


def mssql_query(query):
    query = re.sub('(\n){2,5}', '\n', query)
    return query


def format_correct(str):
    return re.sub('\'', '"', str)


def str_join(a, b):
    return str(a) + str(b)


def build_dict_from_xlsx(dest_file, sheet_name):
    wb = load_workbook(filename = dest_file)
    ws = wb[sheet_name]
    key_words_dict = {}

    index = 1
    while index <= ws.max_row:
        key = normalWord(ws[str_join('A', index)].value)
        value = ws[str_join('B', index)].value
        if key:
            if value:
                key_words_dict[key] = value
            else:
                key_words_dict[key] = ""

        index += 1

    return key_words_dict


def build_tokenize(text):
    result = []
    for para in text.split('\n'):
        sents = sent_tokenize(para)
        result.append([])
        for sent in sents:
            words = word_tokenize(sent)
            result[-1].append(tuple(words))

    return result


def recover_text_from_tokenize(tokenized_text):
    for i, para in enumerate(tokenized_text):
        for j, sent in enumerate(para):
            tokenized_text[i][j] = link_words_to_phrase(sent)
        tokenized_text[i] = link_words_to_phrase(tokenized_text[i]) + '\n'
    text = link_words_to_phrase(tokenized_text)
    return text




def link_words_to_phrase(list):
    if not len(list):
        return ''
    result = " ".join(str(item) for item in list)
    return result


def tokenize_list_replace_keywords(tokenize_sent, keywords_dict):
    queue = tokenize_sent
    result = []
    while len(queue):
        for i in list(range(1, len(queue)+1))[::-1]:
            current_phrase = normalWord(link_words_to_phrase(queue[:i]))
            if current_phrase in keywords_dict:
                result.append(keywords_dict[current_phrase])
                queue = queue[i:]
                break
            elif i == 1:
                result.append(queue[0])
                queue = queue[1:]
    return result


def tokenize_replace_keywords(text, keywords_dict, output= 'text'):
    tokenized_text = build_tokenize(text)
    for i, para in enumerate(tokenized_text):
        for j, sent in enumerate(para):
            tokenized_text[i][j] = tuple(tokenize_list_replace_keywords(sent, keywords_dict))

    if output == 'text':
        return recover_text_from_tokenize(tokenized_text)
    return tokenized_text


def regx_replace_keywords(text, keywords_dict):
    for key, value in keywords_dict.items():
        try:
            text = re.sub(key, value, text, flags=re.IGNORECASE)
        except Exception as e:
            continue
    return text


if __name__ == "__main__":
    ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
    # print(check_contain_chinese('中国天天hasdh'))
    # print(check_contain_chinese('xxx'))
    # print(check_contain_chinese('xx中国'))
    text = 'I love you baby, do you love me?\noh, you will never bechave me, is it? you reall think i am a idoit?, enhancement unbeatable location moments away'
    keywords_dict = build_dict_from_xlsx(os.path.join(ROOT_DIR, 'tmp/SampleEnglishKeyword-20180701.xlsx'), 'English')
    print(build_tokenize(text))
    a = tokenize_replace_keywords(text, keywords_dict, 'text')
    print(a)
    # print(regx_replace_keywords(text, keywords_dict))

    t = ['1', '2', '3', '4', '5']
    d = {'1': 'a', '3 4': 'c d', '2 3': 'qwe'}
    print(tokenize_list_replace_keywords(t, d))


    # a = build_tokenize(text)
    # for i in a:
    #     print(i)





