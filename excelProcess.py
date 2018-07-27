from openpyxl import load_workbook
from openpyxl.styles import Alignment
from lingmoAPI import lingmoAPI
from funcs import str_join, regx_replace_keywords, build_dict_from_xlsx, tokenize_replace_keywords
import os


def translator_xlsx(dest_file, english_keywords_replacement_dict=None, chinese_keywords_replacement_dict=None):

    lw = load_workbook(filename=dest_file)
    ws = lw.active

    api = lingmoAPI()

    index = 5
    total_count = 0
    while index <= ws.max_row:
        value = ws[str_join('B', index)].value
        if value:
            if english_keywords_replacement_dict:
                value = tokenize_replace_keywords(value, english_keywords_replacement_dict)
            translation = api.en_to_ch(value)
            if chinese_keywords_replacement_dict:
                treated_translation = regx_replace_keywords(translation, chinese_keywords_replacement_dict)
                ws[str_join('E', index)] = treated_translation
            else:
                ws[str_join('E', index)] = translation

            ws[str_join('F', index)] = value
            ws[str_join('F', index)].alignment = Alignment(wrapText=True)
            ws[str_join('E', index)].alignment = Alignment(wrapText=True)
            total_count += 1
        index += 1

    print(total_count)

    lw.save(dest_file)



ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
english_keywords_dict = build_dict_from_xlsx(os.path.join(ROOT_DIR, 'tmp/SampleEnglishKeyword-20180703.xlsx'), 'English')
chinese_keywords_dict = build_dict_from_xlsx(os.path.join(ROOT_DIR, 'tmp/SampleEnglishKeyword-20180703.xlsx'), 'Chinese')
# print(chinese_keywords_dict)
translator_xlsx(os.path.join(ROOT_DIR,'tmp', '5.26.xlsx'), english_keywords_replacement_dict=english_keywords_dict, chinese_keywords_replacement_dict=chinese_keywords_dict)






